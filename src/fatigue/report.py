"""Menyusun laporan absensi & fatigue, dan menulisnya ke Excel.

Yang dijawab laporan ini adalah pertanyaan yang benar-benar ditanyakan orang:
siapa hadir, berapa lama ia terpantau, seberapa sering matanya tertutup, dan
kapan persisnya ia sempat berbahaya. Bukan dump tabel mentah.

Agregasi dihitung dari `fatigue_samples` — cuplikan berkala tiap ~30 detik.
Konsekuensinya yang perlu diingat saat membaca angkanya: **lama pengamatan
dihitung dari jumlah cuplikan dikali intervalnya**, bukan dari selisih jam
pertama dan terakhir. Kalau seseorang keluar ruangan satu jam lalu kembali,
jam kerjanya tidak ikut terhitung sebagai "terpantau" — dan itu memang yang
benar, karena selama satu jam itu sistem tidak melihat apa-apa tentang dia.

Rata-rata PERCLOS juga dihitung hanya atas cuplikan yang levelnya bukan
TIDAK_DIKETAHUI (sudah disaring saat penulisan), sehingga menit-menit kalibrasi
awal tidak mengencerkan angkanya.
"""
from __future__ import annotations

import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from src.fatigue.attendance import AttendanceBook
from src.fatigue.records import FatigueLog
from src.fatigue.types import FatigueLevel

# Urutan level dari paling ringan, dipakai untuk kolom "lama di tiap level".
LEVEL_ORDER = [
    FatigueLevel.ALERT,
    FatigueLevel.MILD,
    FatigueLevel.SEVERE,
    FatigueLevel.CRITICAL,
]


def _fmt_duration(seconds: float) -> str:
    """Detik -> "2j 15m" / "15m" / "45d". Lebih terbaca daripada angka detik."""
    seconds = max(0.0, seconds)
    hours, rest = divmod(int(seconds), 3600)
    minutes, secs = divmod(rest, 60)
    if hours:
        return f"{hours}j {minutes}m"
    if minutes:
        return f"{minutes}m"
    return f"{secs}d"


def _clock(ts: float | None) -> str:
    return time.strftime("%H:%M:%S", time.localtime(ts)) if ts else ""


def _date(ts: float | None) -> str:
    return time.strftime("%Y-%m-%d", time.localtime(ts)) if ts else ""


@dataclass
class PersonDay:
    """Ringkasan satu orang pada satu hari."""

    date: str
    employee_id: str
    name: str
    first_seen: float | None = None
    last_seen: float | None = None
    observed_seconds: float = 0.0
    perclos_mean: float = 0.0
    perclos_max: float = 0.0
    blink_mean: float = 0.0
    yawn_total: float = 0.0
    microsleep_max: int = 0
    longest_closure: float = 0.0
    worst_level: FatigueLevel = FatigueLevel.UNKNOWN
    seconds_at: dict[str, float] = field(default_factory=dict)
    alert_count: int = 0
    attendance_time: float | None = None

    def to_row(self) -> dict:
        return {
            "Tanggal": self.date,
            "ID": self.employee_id,
            "Nama": self.name,
            "Absen masuk": _clock(self.attendance_time),
            "Terlihat pertama": _clock(self.first_seen),
            "Terlihat terakhir": _clock(self.last_seen),
            "Lama terpantau": _fmt_duration(self.observed_seconds),
            "PERCLOS rata2": round(self.perclos_mean, 4),
            "PERCLOS tertinggi": round(self.perclos_max, 4),
            "Kedip/menit": round(self.blink_mean, 1),
            "Menguap/menit": round(self.yawn_total, 1),
            "Microsleep terbanyak": self.microsleep_max,
            "Terpejam terlama (dtk)": round(self.longest_closure, 1),
            "Level terburuk": self.worst_level.value,
            **{f"Lama {lv.value}": _fmt_duration(self.seconds_at.get(lv.value, 0.0))
               for lv in LEVEL_ORDER},
            "Jumlah peringatan": self.alert_count,
        }


@dataclass
class Report:
    start: float
    end: float
    generated_at: float
    person_days: list[PersonDay] = field(default_factory=list)
    attendance_rows: list[dict] = field(default_factory=list)
    event_rows: list[dict] = field(default_factory=list)
    employees: list[dict] = field(default_factory=list)
    sample_interval: float = 30.0
    notes: list[str] = field(default_factory=list)

    @property
    def period_label(self) -> str:
        return f"{_date(self.start)} s/d {_date(self.end - 1)}"

    def team_summary(self) -> list[dict]:
        """Rekap seluruh tim untuk periode ini, satu baris per orang."""
        by_person: dict[str, list[PersonDay]] = defaultdict(list)
        for day in self.person_days:
            by_person[f"{day.employee_id}|{day.name}"].append(day)

        rows = []
        for key, days in by_person.items():
            employee_id, name = key.split("|", 1)
            observed = sum(d.observed_seconds for d in days)
            # Rata-rata PERCLOS DITIMBANG lama pengamatan. Merata-ratakan
            # rata-rata harian akan memberi bobot sama pada hari yang cuma
            # terpantau 5 menit dan hari yang terpantau 8 jam.
            perclos = (
                sum(d.perclos_mean * d.observed_seconds for d in days) / observed
                if observed else 0.0
            )
            worst = max((d.worst_level for d in days),
                        key=lambda lv: lv.severity, default=FatigueLevel.UNKNOWN)
            seconds_at: dict[str, float] = defaultdict(float)
            for d in days:
                for level, secs in d.seconds_at.items():
                    seconds_at[level] += secs
            rows.append({
                "ID": employee_id,
                "Nama": name,
                "Hari hadir": len(days),
                "Total terpantau": _fmt_duration(observed),
                "PERCLOS rata2": round(perclos, 4),
                "PERCLOS tertinggi": round(max((d.perclos_max for d in days), default=0.0), 4),
                "Microsleep terbanyak": max((d.microsleep_max for d in days), default=0),
                "Terpejam terlama (dtk)": round(max((d.longest_closure for d in days), default=0.0), 1),
                "Level terburuk": worst.value,
                **{f"Lama {lv.value}": _fmt_duration(seconds_at.get(lv.value, 0.0))
                   for lv in LEVEL_ORDER},
                "Total peringatan": sum(d.alert_count for d in days),
            })
        rows.sort(key=lambda r: (-r["Total peringatan"], -r["PERCLOS rata2"]))
        return rows


def build_report(
    db_path: str | Path | None = None,
    start: float | None = None,
    end: float | None = None,
    sample_interval: float | None = None,
) -> Report:
    """Kumpulkan semuanya dari database untuk rentang [start, end).

    Default: hari ini. `sample_interval` dipakai mengubah jumlah cuplikan jadi
    durasi; kalau tidak diberikan, diambil dari `FatigueLog` yang aktif.
    """
    now = time.time()
    if start is None:
        local = time.localtime(now)
        start = time.mktime((local.tm_year, local.tm_mon, local.tm_mday,
                             0, 0, 0, 0, 0, -1))
    if end is None:
        end = start + 86400

    book = AttendanceBook(db_path=db_path)
    log = FatigueLog(db_path=db_path)
    interval = sample_interval if sample_interval is not None else log.sample_interval

    samples = log.samples(since=start, until=end)
    events = log.events(since=start, until=end)
    attendance = book.records(since=start, limit=100_000)
    attendance = [r for r in attendance if r.timestamp < end]

    # ---- kelompokkan cuplikan per (tanggal, orang) ----
    grouped: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for s in samples:
        key = (_date(s["timestamp"]), s["employee_id"] or "", s["display_name"])
        grouped[key].append(s)

    first_attendance: dict[tuple[str, str], float] = {}
    for record in attendance:
        key = (_date(record.timestamp), record.employee_id)
        if key not in first_attendance or record.timestamp < first_attendance[key]:
            first_attendance[key] = record.timestamp

    alerts_per_day: dict[tuple[str, str], int] = defaultdict(int)
    for event in events:
        alerts_per_day[(_date(event.timestamp), event.employee_id or "")] += 1

    person_days = []
    for (date, employee_id, name), rows in sorted(grouped.items()):
        perclos = [r["perclos"] for r in rows]
        seconds_at: dict[str, float] = defaultdict(float)
        for r in rows:
            seconds_at[r["level"]] += interval

        worst = max(
            (FatigueLevel(r["level"]) for r in rows),
            key=lambda lv: lv.severity, default=FatigueLevel.UNKNOWN,
        )
        person_days.append(PersonDay(
            date=date,
            employee_id=employee_id or "(tidak dikenal)",
            name=name,
            first_seen=min(r["timestamp"] for r in rows),
            last_seen=max(r["timestamp"] for r in rows),
            observed_seconds=len(rows) * interval,
            perclos_mean=sum(perclos) / len(perclos),
            perclos_max=max(perclos),
            blink_mean=sum(r["blink_rate"] for r in rows) / len(rows),
            yawn_total=sum(r["yawn_rate"] for r in rows) / len(rows),
            microsleep_max=max(r["microsleep_count"] for r in rows),
            longest_closure=max(r["longest_closure"] for r in rows),
            worst_level=worst,
            seconds_at=dict(seconds_at),
            alert_count=alerts_per_day.get((date, employee_id), 0),
            attendance_time=first_attendance.get((date, employee_id)),
        ))

    notes = []
    if not samples:
        notes.append(
            "Tidak ada cuplikan fatigue pada rentang ini. Data fatigue baru "
            "tercatat saat sesi monitoring berjalan (CLI webcam atau tab "
            "Monitor di Streamlit)."
        )
    unknown = {d.name for d in person_days if d.employee_id == "(tidak dikenal)"}
    if unknown:
        notes.append(
            f"{len(unknown)} orang terpantau tanpa identitas — mereka belum "
            "terdaftar di absensi, jadi hanya muncul sebagai baris terpisah "
            "tanpa ID."
        )

    return Report(
        start=start, end=end, generated_at=now,
        person_days=person_days,
        attendance_rows=[r.to_row() for r in attendance],
        event_rows=[e.to_row() for e in events],
        employees=[e.to_dict() for e in book.list_employees()],
        sample_interval=interval,
        notes=notes,
    )


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------
def to_excel(report: Report, path: str | Path) -> Path:
    """Tulis laporan ke satu file .xlsx berisi beberapa sheet.

    Format dijaga tetap sederhana dan bisa di-pivot: satu baris = satu
    pengamatan, header di baris pertama, tanpa sel yang digabung. Laporan yang
    cantik tapi tidak bisa di-pivot akan langsung diketik ulang oleh orang
    yang menerimanya.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")

    def add_sheet(title: str, rows: list[dict], columns: list[str] | None = None):
        ws = wb.create_sheet(title[:31])
        if not rows:
            ws["A1"] = "(tidak ada data pada rentang ini)"
            return ws
        columns = columns or list(rows[0].keys())
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in rows:
            ws.append([row.get(c, "") for c in columns])

        # Lebar kolom mengikuti isi terpanjang, dibatasi supaya kolom "alasan"
        # yang panjang tidak membuat sheet-nya mustahil dibaca.
        for i, column in enumerate(columns, start=1):
            longest = max([len(str(column))] + [len(str(r.get(column, ""))) for r in rows])
            ws.column_dimensions[get_column_letter(i)].width = min(46, max(10, longest + 2))
        ws.freeze_panes = "A2"
        # Autofilter membuat penerima bisa langsung menyaring per nama/tanggal.
        ws.auto_filter.ref = ws.dimensions
        return ws

    # ---- Ringkasan ----
    info = wb.active
    info.title = "Ringkasan"
    info["A1"] = "Laporan Absensi & Fatigue"
    info["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Periode", report.period_label),
        ("Dibuat", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(report.generated_at))),
        ("Interval cuplikan", f"{report.sample_interval:.0f} detik"),
        ("Karyawan terdaftar", len(report.employees)),
        ("Baris absensi", len(report.attendance_rows)),
        ("Kejadian fatigue", len(report.event_rows)),
        ("Orang-hari terpantau", len(report.person_days)),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        info[f"A{i}"] = label
        info[f"A{i}"].font = Font(bold=True)
        info[f"B{i}"] = value
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 40

    row = 3 + len(meta) + 1
    info[f"A{row}"] = "Cara membaca"
    info[f"A{row}"].font = Font(bold=True)
    for i, line in enumerate([
        "Lama terpantau = jumlah cuplikan x interval, bukan selisih jam pertama "
        "dan terakhir. Waktu saat orangnya tidak terlihat kamera tidak dihitung.",
        "PERCLOS = fraksi waktu kelopak mata tertutup. Orang terjaga normal "
        "berada di sekitar 3-5%.",
        "Microsleep = mata terpejam >= 1,5 detik terus-menerus.",
        "Level: SEGAR < WASPADA < LELAH < KRITIS. TIDAK_DIKETAHUI tidak ikut "
        "dicuplik dan tidak muncul di laporan ini.",
        *report.notes,
    ], start=1):
        info[f"A{row + i}"] = line
        info[f"A{row + i}"].alignment = Alignment(wrap_text=True, vertical="top")
        info.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=8)
        info.row_dimensions[row + i].height = 30

    add_sheet("Rekap per orang", report.team_summary())
    add_sheet("Harian per orang", [d.to_row() for d in report.person_days])
    add_sheet("Log absensi", report.attendance_rows)
    add_sheet("Kejadian fatigue", report.event_rows)
    add_sheet("Karyawan", [{
        "ID": e["employee_id"], "Nama": e["name"], "Departemen": e["department"],
        "Foto wajah": e["num_embeddings"], "Aktif": "ya" if e["active"] else "tidak",
    } for e in report.employees])

    wb.save(path)
    return path


def to_excel_bytes(report: Report) -> bytes:
    """Versi in-memory, untuk tombol unduh Streamlit."""
    import io
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        path = to_excel(report, Path(tmp) / "laporan.xlsx")
        return io.BytesIO(path.read_bytes()).getvalue()
